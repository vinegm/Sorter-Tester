import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import copy
import time
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path


def calculate_average_time(array_times: list) -> int:
    """Takes a array with times samples and calculates the average of them
    
    Parameters:
    array_times(np.array): Array containing the samples

    Returns:
    average_time(int): Time in average the algorithm took in microseconds
    """
    try:
        sum_times = array_times.sum()

    except AttributeError:
        sum_times = sum(array_times)

    # Calculates the average time 1.000.000 to turn the value into a int
    average_time = int((sum_times / len(array_times))*1_000_000)

    return average_time


def test_timing(test_case: list, sorter, amount_tests: int) -> int:
    """Tests a sorter a given amount of times to get a average time
    
    Parameters:
    test_case(list/np.array): Array used for the tests
    sorter: Sorter used for the tests
    amount_tests(int): How many tests will be made to get a average time

    Returns:
    average_time(int): Time the sorter took in average to complete the task
    """
    if amount_tests < 1:
        raise ValueError("Amount of tests must be 1 or more!")
    
    # Creates an empty array to store the results of every test
    test_times = np.array([])

    tests_made = 0
    while (tests_made < amount_tests):
        # Makes a copy of the test case to not interfere in the next tests
        case = np.copy(test_case)

        # Times the sorter doing its task
        startTimer = time.perf_counter()
        sorter(case)
        stopTimer = time.perf_counter()

        if tests_made == 0 and amount_tests > 1:
            if (est_time := stopTimer - startTimer) > 0.01:
                print(f"(Estimated time in seconds: {(est_time * amount_tests):.2f})")
            else:
                print()

        test_times = np.append(test_times, (stopTimer - startTimer))

        tests_made += 1

    # Calculates the average time across all the tests
    average_time = calculate_average_time(test_times)

    return average_time


def test_sorter(test_case: list, sorter, amount_tests: int) -> tuple:
    """Tests a given sorter
    
    Parameters:
    test_case(array): Test that will be run
    sorter(list): Sorter that will be tested
    amount_tests(int): How many tests will be made to get a average time for sorting

    Returns:
    average_time: Average time of the sorter to complete the task on the given test
    comparisons: How many comparisons the sorter made
    swaps: How many comparisons the sorter swaps
    """
    if amount_tests < 1:
        raise ValueError("Amount of tests must be 1 or more!")
    
    # Makes a copy of the case, so one test doesn't effect the other
    case = np.copy(test_case)

    # Calculates the average time of the sorter
    average_time = test_timing(case, sorter[0], amount_tests)

    # If doesn't have a sorter for counting comparisons/swaps, sets them to 0
    if sorter[1] == None:
        comparisons = 0
        swaps = 0
        return average_time, comparisons, swaps

    # Tests amount of comparisons and swaps
    sorted_array, comparisons, swaps = sorter[1](case)

    if not np.array_equal(sorted_array, sorted(test_case)):
        print(sorted_array)
        raise Exception(f"Sorter Not Working!")
    
    return average_time, comparisons, swaps


def test_sorters(test_cases: dict, sorters, amount_tests: list) -> list:
    """Tests the average time, amount of comparisons and swaps for multiple sorters in multiple tests
    
    Parameters:
    test_cases(list) = list of cases to be tested on
    sorters(list) = List of sorters to be tested
    amount_tests(int) = Amount of tests that will be run

    Returns:
    cases(list) = List of data frames for each test case, containing the results for every sorter 
    """

    if amount_tests < 1:
        raise ValueError("Amount of tests must be 1 or more!")

    # Creates a dict for storing the results
    data = {"sorter": [],
            "average time": [],
            "comparisons": [],
            "swaps": []}
    results = {}

    print("-" * 50)
    for case_name, case in test_cases.items():
        print(f"{case_name}:")
        case_results = copy.deepcopy(data)
        for sorter in sorters:

            if amount_tests > 1:
                print(f"Testing {sorter[0].__name__}... ", end = "")
            else:
                print(f"Testing {sorter[0].__name__}...")

            # Test the one sorter at a time in a case
            average_time, comparisons, swaps = test_sorter(case, sorter, amount_tests)
            
            # Stores the results
            case_results["sorter"].append(sorter[0].__name__)
            case_results["average time"].append(average_time)
            case_results["comparisons"].append(comparisons)
            case_results["swaps"].append(swaps)
        
        # Turns the dict into a data frame
        results[case_name] = pd.DataFrame(case_results)
        print("-" * 50)

    return results


def directory_path(cases_ran: dict, amount_tests: int) -> Path:
    """Returns the directory path with a readme for context of the text"""

    results_dir = Path("results")

    directory = results_dir/"test"

    i = 0
    while directory.exists():
        i += 1
        directory = results_dir / f"test ({i})"

    directory.mkdir(parents=True)
    (directory/"graphs").mkdir(parents=True)

    with open(directory/"README.txt", "w") as file:
        file.write(f"This this test ran {amount_tests} times to get an average time\
                   \nIt ran the cases:\n{cases_ran}")

    return directory


def save_dataframes(directory: Path, cases: dict) -> None:
    """Saves the results to a excel

    Parameters:
    directory(Path): Where the results will be saved
    cases(dict): Dict of cases that will be saved in a excel
    """
    workbook = Workbook()

    # Remove the default sheet openpyxl creates
    workbook.remove(workbook['Sheet'])

    # Settings of the fonts and cells
    font = Font(size = 14)
    border = Border(left = Side(style = "thin"),
                    right = Side(style = "thin"),
                    top = Side(style = "thin"),
                    bottom = Side(style = "thin"))
    gray_cell = PatternFill(start_color = "c4c4c4", fill_type = "solid")
    
    # Iterates every case, creating a sheet for each
    for case, results in cases.items():
        sheet = workbook.create_sheet(title = case)

        rows = dataframe_to_rows(results, index = False)

        # Populates the sheet
        for row_i, row in enumerate(rows, 1):
            for col_i, value in enumerate(row, 1):
                sheet.cell(row = row_i, column = col_i, value = value)

        # Iterates every column
        for i, column in enumerate(sheet.columns, 1):
            # For even columns, sets their colors to gray
            if i % 2 == 0:
                for cell in column:
                    cell.fill = gray_cell

            # Looks for the cell with the biggest lenght
            value_length = 0
            for cell in column:
                if (cell_length := len(str(cell.value))) > value_length:
                    value_length = cell_length

                # Applies the font and border to each cell
                cell.font = font
                cell.border = border

            # Adjusts the width of the cells in the column
            adjusted_length = (value_length + 1) * 1.4
            column_letter = get_column_letter(column[0].column)
            sheet.column_dimensions[column_letter].width = adjusted_length

    file = directory/"Results.xlsx"
    workbook.save(file)

    workbook.close()    


def save_graphs(directory: Path, cases: dict) -> None:
    """Saves a png for each case

    Parameters:
    cases(dict): Dict of cases that will be saved
    """
    for case, results in cases.items():
        fig, axs = plt.subplots(3, 1, figsize = (10, 10))
        
        axs[0].bar(results["sorter"], results["average time"], color = "y")
        axs[0].set_title(case)
        axs[0].set_ylabel("Time (microseconds)")
        
        axs[1].bar(results["sorter"], results["comparisons"], color = "b")
        axs[1].set_ylabel("Comparisons")
    
        axs[2].bar(results["sorter"], results["swaps"], color = "g")
        axs[2].set_xlabel("Sorter")
        axs[2].set_ylabel("Swaps")
        
        plt.tight_layout()
    
        plt.savefig(f"{directory}/graphs/{case}.png")
